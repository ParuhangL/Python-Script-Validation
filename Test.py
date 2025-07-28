import pandas as pd
import tkinter as tk
from tkinter.filedialog import askopenfilenames
import tkinter.messagebox as msgbox
from nepali_datetime import date as NepaliDate
import os
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === GUI ===
root = tk.Tk()
root.withdraw()
input_paths = askopenfilenames(
    title="Select Excel files for validation",
    filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
)
if not input_paths:
    exit()

# === Colors for openpyxl ===
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
blue_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

# === RGB for xlwings ===
RED_RGB = (255, 199, 206)
BLUE_RGB = (187, 222, 251)
GREEN_RGB = (200, 230, 201)
ORANGE_RGB = (255, 224, 178)

current_bs = NepaliDate.today()

def is_effectively_blank(val):
    if pd.isna(val):
        return True
    try:
        s = str(val).strip()
        return s == "" or s.lower() in ["''", '=""', 'nan']
    except:
        return True

def normalize_date(val):
    val_str = str(val).strip() if pd.notna(val) else ""
    parts = val_str.split('.')
    if len(parts) == 3:
        try:
            y, m, d = map(int, parts)
            if 1900 <= y <= 2200 and 1 <= m <= 12 and 1 <= d <= 32:
                return f"{y}.{m:02d}.{d:02d}", y, m, d, False
        except:
            pass
    return val_str, None, None, None, True

def validate_common_fields(df, cell_func, non_empty_rows=None, is_openpyxl=True):
    if non_empty_rows is None:
        non_empty_rows = ~df.apply(lambda row: all(is_effectively_blank(v) for v in row), axis=1)

    def color(r, c, fill):
        if is_openpyxl:
            cell_func(r, c).fill = fill
        else:
            cell_func(r, c).color = fill

    for col in ['AccountNo', 'LoanAccountNo']:
        if col in df.columns:
            duplicates = df.duplicated(subset=[col], keep=False)
            for i in df.index[non_empty_rows]:
                val = df.at[i, col]
                if pd.isna(val) or str(val).strip() == "":
                    color(i + 2, df.columns.get_loc(col) + 1, red_fill if is_openpyxl else RED_RGB)
                elif duplicates[i]:
                    color(i + 2, df.columns.get_loc(col) + 1, green_fill if is_openpyxl else GREEN_RGB)

    for col in df.columns:
        if 'loanissuedate' in col.lower() and 'bs' in col.lower():
            for i in df.index[non_empty_rows]:
                r, c = i + 2, df.columns.get_loc(col) + 1
                norm, y, m, d, err = normalize_date(df.at[i, col])
                if is_openpyxl:
                    cell_func(r, c).value = norm
                if err:
                    color(r, c, red_fill if is_openpyxl else RED_RGB)
                else:
                    try:
                        if NepaliDate(y, m, d) > current_bs:
                            color(r, c, red_fill if is_openpyxl else RED_RGB)
                    except:
                            color(r, c, red_fill if is_openpyxl else RED_RGB)

    for col in ['DeposittypeCode', 'InterestRate', 'LoanTypeCode']:
        if col in df.columns:
            for i in df.index[non_empty_rows]:
                val = df.at[i, col]
                if pd.isna(val) or str(val).strip() == "":
                    color(i + 2, df.columns.get_loc(col) + 1, red_fill if is_openpyxl else RED_RGB)

    for col in ['PeriodType', 'DurationType']:
        if col in df.columns:
            for i in df.index[non_empty_rows]:
                val = df.at[i, col]
                if pd.isna(val) or len(str(val).strip()) != 1:
                    color(i + 2, df.columns.get_loc(col) + 1, red_fill if is_openpyxl else RED_RGB)

def validate_xlsx_xlsm(path):
    wb = load_workbook(path, keep_vba=path.lower().endswith('.xlsm'))
    ws = wb.active
    # Normalize columns to strip and remove non-breaking spaces
    headers = [cell.value.strip() if isinstance(cell.value, str) else str(cell.value).strip()
               for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    df = pd.DataFrame([[cell.value for cell in row] for row in ws.iter_rows(min_row=2)], columns=headers)

    # Normalize dataframe columns similarly for matching
    df.columns = [str(c).strip().replace('\u00A0', '') for c in df.columns]

    def cell(r, c): return ws.cell(row=r, column=c)

    non_empty_rows = ~df.apply(lambda row: all(is_effectively_blank(v) for v in row), axis=1)

    # === Member ID and Name Check ===
    id_cols = [col for col in df.columns if any(k in col.lower() for k in ['id', 's#', 'n#'])]
    name_cols = [col for col in df.columns if 'name' in col.lower()]

    if id_cols and name_cols:
        id_col = id_cols[0]
        name_col = name_cols[0]
        id_idx = df.columns.get_loc(id_col) + 1
        name_idx = df.columns.get_loc(name_col) + 1

        df[id_col] = df[id_col].fillna('').astype(str).str.replace(r'\s+', '', regex=True)
        df[name_col] = df[name_col].fillna('').astype(str).str.replace(r'\s+', '', regex=True)

        both_blank = (df[id_col] == "") & (df[name_col] == "")
        check_rows = non_empty_rows & ~both_blank

        duplicate_mask = df.duplicated(subset=[id_col], keep=False) & check_rows

        for i in df.index[check_rows]:
            r = i + 2
            val = df.at[i, id_col]
            if not val or val.lower() == 'nan':
                ws.cell(row=r, column=id_idx).fill = red_fill
            elif duplicate_mask[i]:
                ws.cell(row=r, column=id_idx).fill = green_fill

            val_name = df.at[i, name_col]
            if not val_name or val_name.lower() == 'nan':
                ws.cell(row=r, column=name_idx).fill = red_fill

        valid_mask = check_rows & ~duplicate_mask & df[id_col].apply(lambda x: x.isdigit())
        df.loc[valid_mask, id_col] = range(1, valid_mask.sum() + 1)
        for i in df.index[check_rows]:
            ws.cell(row=i + 2, column=id_idx).value = df.at[i, id_col]

    # === Date Fields ===
    print(f"All columns: {list(df.columns)}")
    date_cols = [col for col in df.columns if any(k in col.lower() for k in ["onbs", "datebs", "issuedate", "maturity", "birth", "open"])]
    print(f"Filtered date columns (containing 'onbs'): {date_cols}")

    for date_col in date_cols:
        col_i = df.columns.get_loc(date_col) + 1
        col_lower = date_col.lower().replace(" ", "").replace("\u00A0", "")
        for i in df.index[non_empty_rows]:
            r = i + 2
            val = df.at[i, date_col]
            norm, y, m, d, err = normalize_date(val)
            ws.cell(row=r, column=col_i).value = norm

            if err:
                fill = orange_fill if "birth" in col_lower else red_fill
                ws.cell(row=r, column=col_i).fill = fill
                continue

            if "birth" in col_lower:
                try:
                    birth_date = NepaliDate(y, m, d)
                    if birth_date > current_bs:
                        ws.cell(row=r, column=col_i).fill = orange_fill
                except:
                    ws.cell(row=r, column=col_i).fill = orange_fill

            elif "accountopen" in col_lower:
                try:
                    if NepaliDate(y, m, d) > current_bs:
                        ws.cell(row=r, column=col_i).fill = red_fill
                except:
                    ws.cell(row=r, column=col_i).fill = red_fill

            elif "maturity" in col_lower:
                print(f"Validating maturity column: '{date_col}'")
                open_col = next((c for c in df.columns if "accountopen" in c.lower()), None)
                loan_col = next((c for c in df.columns if "loanissuedate" in c.lower()), None)
                reference_col = open_col if open_col else loan_col
                if reference_col:
                    ref_val = df.at[i, reference_col]
                    ref_norm, ry, rm, rd, ref_err = normalize_date(ref_val)
                    if not ref_err:
                        try:
                            maturity_date = NepaliDate(y, m, d)
                            ref_date = NepaliDate(ry, rm, rd)
                            if maturity_date < ref_date:
                                ws.cell(row=r, column=col_i).fill = red_fill
                        except:
                            ws.cell(row=r, column=col_i).fill = red_fill
                    else:
                        ws.cell(row=r, column=col_i).fill = red_fill

    # === ClosingBalance Null Check ===
    if 'ClosingBalance' in df.columns:
        col_i = df.columns.get_loc('ClosingBalance') + 1
        for i in df.index[non_empty_rows]:
            val = df.at[i, 'ClosingBalance']
            if pd.isna(val) or str(val).strip() == "":
                ws.cell(row=i + 2, column=col_i).fill = red_fill

    # === Balance Checks ===
    for balance_col in [col for col in df.columns if any(k in col.lower() for k in ['balance', 'amount'])]:
        if any(ex in balance_col.lower() for ex in ['installment', 'payable']):
            continue
        col_i = df.columns.get_loc(balance_col) + 1
        for i in df.index[non_empty_rows]:
            val = df.at[i, balance_col]
            try:
                if float(val) <= 0:
                    ws.cell(row=i + 2, column=col_i).fill = blue_fill
            except:
                ws.cell(row=i + 2, column=col_i).fill = blue_fill

    # === Share Amount Validation ===
    for share_col in [col for col in df.columns if "share" in col.lower()]:
        col_i = df.columns.get_loc(share_col) + 1
        for i in df.index[non_empty_rows]:
            r = i + 2
            val = df.at[i, share_col]
            try:
                if pd.isna(val) or int(float(val)) % 100 != 0:
                    ws.cell(row=r, column=col_i).fill = red_fill
            except:
                ws.cell(row=r, column=col_i).fill = red_fill

    # === Validate common fields ===
    validate_common_fields(df, cell, non_empty_rows=non_empty_rows, is_openpyxl=True)

    wb.save(path)
    print(f"Processed: {os.path.basename(path)}")

    os.startfile(path)


def validate_xls(path):
    app = xw.App(visible=False)
    wb = app.books.open(path)
    sht = wb.sheets[0]
    df = sht.used_range.options(pd.DataFrame, index=False, header=1).value

    def cell(r, c): return sht.cells(r, c)

    non_empty_rows = ~df.apply(lambda row: all(is_effectively_blank(v) for v in row), axis=1)
    validate_common_fields(df, cell, non_empty_rows=non_empty_rows, is_openpyxl=False)

    wb.save()
    wb.close()
    app.quit()
    print(f"Processed: {os.path.basename(path)}")

    os.startfile(path)


for path in input_paths:
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == '.xls':
            validate_xls(path)
        else:
            validate_xlsx_xlsm(path)
    except PermissionError:
        msgbox.showwarning("File Save Error", f"Please close '{os.path.basename(path)}' and try again.")
    except Exception as e:
        msgbox.showerror("Unexpected Error", f"{os.path.basename(path)}:\n{str(e)}")
